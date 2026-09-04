"""
OpenTargets Association Score Enrichment Script
================================================
Reads an Excel file with 'indication' and 'moa' columns,
maps MoA entries and indications to OpenTargets equivalents using Gemini + Google Search,
fetches association scores from OpenTargets GraphQL API,
and writes the highest association score per indication back to the Excel file.

Features
--------
- Cache: resolved MoAs and indications are saved to a JSON file next to the
  input file (<stem>_cache.json) and reused on subsequent runs.
- Batching: each Gemini call resolves 3 indications at once.
- Parallelism: 3 concurrent workers for both resolution and scoring.
- Retry: each LLM call retries once on failure.

Usage:
    pip install -r requirements.txt
    python opentargets_association.py

Environment variables (in .env):
    GEMINI_API_KEY   - Google Gemini API key
    GEMINI_MODEL     - Model name (default: gemini-2.5-flash)
    INPUT_FILE       - Path to the input Excel file
    OUTPUT_FILE      - (Optional) output path; defaults to <input>_scored.xlsx
    SHEET_NAME       - (Optional) sheet name; defaults to first sheet
    CACHE_FILE       - (Optional) path to cache JSON; defaults to <input>_cache.json
    WORKERS          - (Optional) parallel workers (default: 3)
"""

import os
import re
import sys
import time
import json
import logging
import threading
import requests
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor, as_completed

# ── Load .env ──────────────────────────────────────────────────────────────────
load_dotenv()

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
GEMINI_MODEL   = os.getenv("GEMINI_MODEL", "gemini-2.5-flash").strip()
INPUT_FILE     = os.getenv("INPUT_FILE", "").strip()
OUTPUT_FILE    = os.getenv("OUTPUT_FILE", "").strip()
SHEET_NAME     = os.getenv("SHEET_NAME", "").strip() or None
CACHE_FILE     = os.getenv("CACHE_FILE", "").strip()
WORKERS        = int(os.getenv("WORKERS", "3"))

if not GEMINI_API_KEY:
    sys.exit("❌  GEMINI_API_KEY is missing from .env")
if not INPUT_FILE:
    sys.exit("❌  INPUT_FILE is missing from .env")
if not Path(INPUT_FILE).exists():
    sys.exit(f"❌  File not found: {INPUT_FILE}")

_inp = Path(INPUT_FILE)
if not OUTPUT_FILE:
    OUTPUT_FILE = str(_inp.parent / f"{_inp.stem}_scored{_inp.suffix}")
if not CACHE_FILE:
    CACHE_FILE = str(_inp.parent / f"{_inp.stem}_cache.json")

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  [%(threadName)s]  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────
GEMINI_URL   = (
    f"https://generativelanguage.googleapis.com/v1beta/models/"
    f"{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"
)
OT_GRAPHQL   = "https://api.platform.opentargets.org/api/v4/graphql"
HEADERS_JSON = {"Content-Type": "application/json"}
BATCH_SIZE   = 3          # indications per Gemini call
_cache_lock  = threading.Lock()   # protects cache writes

# ── Platform-matching datasource weights ───────────────────────────────────────
# The associatedTargets endpoint computes scores "on the fly" using datasource
# weights. The API defaults differ from the Platform UI defaults (see OT Community
# thread on Release 25.12 score discrepancy). To match Platform scores, we must
# explicitly pass the weights the Platform UI uses. These are extracted from the
# Platform UI's "Export → API Query" feature.
# Set propagate=True for all, weight=1.0 for all, required=False for all.
OT_DATASOURCE_WEIGHTS = [
    {"id": "ot_genetics_portal",     "weight": 1, "propagate": True},
    {"id": "gwas_credible_sets",     "weight": 1, "propagate": True},
    {"id": "eva",                    "weight": 1, "propagate": True},
    {"id": "eva_somatic",            "weight": 1, "propagate": True},
    {"id": "gene_burden",            "weight": 1, "propagate": True},
    {"id": "genomics_england",       "weight": 1, "propagate": True},
    {"id": "gene2phenotype",         "weight": 1, "propagate": True},
    {"id": "uniprot_literature",     "weight": 1, "propagate": True},
    {"id": "uniprot_variants",       "weight": 1, "propagate": True},
    {"id": "orphanet",               "weight": 1, "propagate": True},
    {"id": "clingen",                "weight": 1, "propagate": True},
    {"id": "cancer_gene_census",     "weight": 1, "propagate": True},
    {"id": "intogen",                "weight": 1, "propagate": True},
    {"id": "cancer_biomarkers",      "weight": 1, "propagate": True},
    {"id": "chembl",                 "weight": 1, "propagate": True},
    {"id": "crispr",                 "weight": 1, "propagate": True},
    {"id": "slapenrich",             "weight": 1, "propagate": True},
    {"id": "progeny",                "weight": 1, "propagate": True},
    {"id": "reactome",               "weight": 1, "propagate": True},
    {"id": "sysbio",                 "weight": 1, "propagate": True},
    {"id": "expression_atlas",       "weight": 1, "propagate": True},
    {"id": "europepmc",              "weight": 1, "propagate": True},
    {"id": "impc",                   "weight": 1, "propagate": True},
]

# ── Cache helpers ──────────────────────────────────────────────────────────────

def load_cache() -> dict:
    """Load cache from disk, returning empty structure on first run."""
    if Path(CACHE_FILE).exists():
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
            log.info("📦 Loaded cache from %s  (moas=%d, indications=%d)",
                     CACHE_FILE,
                     len(cache.get("moa_targets", {})),
                     len(cache.get("indications", {})))
            return cache
        except Exception as exc:
            log.warning("⚠️  Cache read failed (%s), starting fresh.", exc)
    return {"moa_targets": {}, "indications": {}}


def save_cache(cache: dict) -> None:
    """Persist cache to disk (thread-safe)."""
    with _cache_lock:
        try:
            with open(CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(cache, f, indent=2, ensure_ascii=False)
        except Exception as exc:
            log.warning("⚠️  Cache write failed: %s", exc)


# ── Gemini helper ──────────────────────────────────────────────────────────────

def gemini_search(prompt: str, max_retries: int = 2) -> str:
    """
    Call Gemini with Google Search grounding.
    Retries once on failure (max_retries=2 means 1 retry).
    """
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {"temperature": 0.1, "maxOutputTokens": 1024},
    }
    for attempt in range(1, max_retries + 1):
        try:
            r = requests.post(GEMINI_URL, json=payload, headers=HEADERS_JSON, timeout=45)
            r.raise_for_status()
            data  = r.json()
            parts = (
                data.get("candidates", [{}])[0]
                    .get("content", {})
                    .get("parts", [])
            )
            return " ".join(p.get("text", "") for p in parts).strip()
        except Exception as exc:
            log.warning("Gemini attempt %d/%d failed: %s", attempt, max_retries, exc)
            if attempt < max_retries:
                time.sleep(2 ** attempt)
    return ""


def extract_id_from_text(text: str, kind: str) -> str | None:
    """Extract ENSG ID (kind='target') or EFO/MONDO/HP/DOID ID (kind='disease')."""
    if kind == "target":
        m = re.search(r"ENSG\d{11}", text)
        return m.group() if m else None
    m = re.search(
        r"(EFO_\d+|MONDO_\d+|HP_\d+|DOID[_:]\d+|Orphanet_\d+|CHEBI_\d+)",
        text, re.IGNORECASE,
    )
    return m.group().replace(":", "_") if m else None


# ── OpenTargets helpers ────────────────────────────────────────────────────────

def _ot_post(query: str, variables: dict, context: str = "") -> dict | None:
    """Shared POST to OT GraphQL with unified error handling."""
    try:
        r = requests.post(
            OT_GRAPHQL,
            json={"query": query, "variables": variables},
            headers=HEADERS_JSON,
            timeout=20,
        )
        if not r.ok:
            log.warning(
                "OT HTTP %s [%s]: %s",
                r.status_code, context, r.text[:400]
            )
            return None
        body = r.json()
        if "errors" in body:
            log.warning("OT GraphQL error [%s]: %s", context, body["errors"])
            return None
        return body.get("data")
    except Exception as exc:
        log.warning("OT request failed [%s]: %s", context, exc)
        return None


def ot_search_target(name: str) -> tuple[str | None, str | None]:
    query = """
    query SearchTarget($q: String!) {
      search(queryString: $q, entityNames: ["target"], page: {index: 0, size: 3}) {
        hits {
          id
          object { ... on Target { approvedSymbol approvedName } }
        }
      }
    }
    """
    data = _ot_post(query, {"q": name}, context=f"target:{name}")
    if data:
        hits = data.get("search", {}).get("hits", [])
        if hits:
            h = hits[0]
            return h["id"], h["object"].get("approvedSymbol")
    return None, None


def ot_search_disease(name: str) -> tuple[str | None, str | None]:
    query = """
    query SearchDisease($q: String!) {
      search(queryString: $q, entityNames: ["disease"], page: {index: 0, size: 3}) {
        hits {
          id
          object { ... on Disease { name } }
        }
      }
    }
    """
    data = _ot_post(query, {"q": name}, context=f"disease:{name}")
    if data:
        hits = data.get("search", {}).get("hits", [])
        if hits:
            h = hits[0]
            return h["id"], h["object"].get("name")
    return None, None


def ot_association_score(disease_id: str, target_id: str) -> float | None:
    """
    Return the overall association score (0-1) for a specific target+disease pair.

    Strategy: use Bs (target ID list filter) to request only the row for our
    target ID, avoiding pagination issues with the previous top-N scan approach.
    Passes datasources explicitly to match Platform UI weights.
    """
    query = """
    query AssocScore($diseaseId: String!, $targetIds: [String!]!,
                     $datasources: [DatasourceSettingsInput!]) {
      disease(efoId: $diseaseId) {
        associatedTargets(
          enableIndirect: false
          Bs: $targetIds
          datasources: $datasources
          page: { index: 0, size: 1 }
        ) {
          rows {
            score
            target { id approvedSymbol }
          }
        }
      }
    }
    """
    data = _ot_post(
        query,
        {"diseaseId": disease_id, "targetIds": [target_id],
         "datasources": OT_DATASOURCE_WEIGHTS},
        context=f"score:{target_id}×{disease_id}",
    )
    if data:
        rows = (
            data.get("disease", {})
                .get("associatedTargets", {})
                .get("rows", [])
        )
        if rows:
            return float(rows[0]["score"])
    return None


# ── Resolution: MoA targets ────────────────────────────────────────────────────

def _resolve_single_moa(moa: str) -> tuple[str, str | None, str | None]:
    """Resolve one MoA string → (moa, ensembl_id, symbol). Retries once."""
    log.info("  Resolving MoA → '%s'", moa)
    prompt = (
        f"Search OpenTargets (platform.opentargets.org) for the target that best "
        f"represents the mechanism of action '{moa}'. "
        f"Return ONLY the Ensembl gene ID (ENSG…) and the approved gene symbol, "
        f"separated by a space. Pick the most common primary target."
    )
    text = gemini_search(prompt)
    ensg = extract_id_from_text(text, "target")
    symbol = None
    if ensg:
        _, symbol = ot_search_target(ensg)
    if not ensg:
        ensg, symbol = ot_search_target(moa)
    if ensg:
        log.info("    ✅ MoA '%s' → %s (%s)", moa, ensg, symbol)
    else:
        log.warning("    ⚠️  Could not resolve target for '%s'", moa)
    return moa, ensg, symbol


def resolve_moa_targets(raw_moas: list[str], cache: dict) -> dict[str, tuple[str | None, str | None]]:
    """
    Resolve MoA strings -> OT targets.
    - Cache hit with a valid ID: use it directly.
    - Cache hit with null ID OR cache miss: fall back to ot_search_target(),
      a direct OT search API call — no Gemini. Updates cache on the fly.
    OpenTargets scores are always fetched fresh on every run.
    """
    target_map: dict[str, tuple[str | None, str | None]] = {}
    cached = cache.get("moa_targets", {})
    cache_updated = False

    for moa in raw_moas:
        entry = cached.get(moa)
        cached_id = entry.get("id") if entry else None

        if entry and cached_id:
            target_map[moa] = (cached_id, entry.get("symbol"))
            log.info("  💾 Cache hit MoA '%s' → %s (%s)", moa, cached_id, entry.get("symbol"))
        else:
            if entry:
                log.warning("  ⚠️  Cache has null ID for MoA '%s' — querying OT search API", moa)
            else:
                log.warning("  ⚠️  MoA '%s' not in cache — querying OT search API", moa)

            # Try gene/target search terms extracted from the MoA description
            ot_id, ot_sym = None, None
            search_terms = _extract_gene_search_terms(moa)
            # Also try the raw MoA string as last resort
            if moa not in search_terms:
                search_terms.append(moa)
            for term in search_terms:
                ot_id, ot_sym = ot_search_target(term)
                if ot_id:
                    log.info("    ✅ OT search resolved MoA '%s' via '%s' → %s (%s)", moa, term, ot_id, ot_sym)
                    break
                log.warning("    ↩ Term '%s' → no match", term)

            target_map[moa] = (ot_id, ot_sym)
            cached[moa] = {"id": ot_id, "symbol": ot_sym}
            cache_updated = True

            if not ot_id:
                log.warning("    ❌ OT search could not resolve MoA '%s'", moa)

    if cache_updated:
        cache["moa_targets"] = cached
        save_cache(cache)

    return target_map


# ── Resolution: Indications (batched) ─────────────────────────────────────────

def _clean_and_parse_json(text: str) -> list[dict]:
    """
    Robustly extract a JSON array from Gemini's free-text response.

    Handles (in order of attempt):
      1. Strip markdown fences and surrounding prose, then json.loads
      2. Find the outermost [...] via bracket-counting (not regex, handles nesting)
      3. Remove trailing commas and retry json.loads
      4. Object-by-object extraction via {…} regex — handles newline-separated objects
      5. Key-value regex extraction per object as last resort
    """
    # ── Step 1: strip markdown fences ─────────────────────────────────────────
    cleaned = re.sub(r"```(?:json|JSON)?\s*", "", text)
    cleaned = re.sub(r"```", "", cleaned).strip()

    # ── Step 2: bracket-counting extraction of outermost [...] ────────────────
    def _extract_array(s: str) -> str | None:
        start = s.find("[")
        if start == -1:
            return None
        depth = 0
        for i, ch in enumerate(s[start:], start):
            if ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    return s[start: i + 1]
        return None  # unbalanced

    array_str = _extract_array(cleaned)
    if array_str:
        # ── Step 3: fix trailing commas, then parse ────────────────────────
        fixed = re.sub(r",\s*([\]\}])", r"\1", array_str)
        try:
            result = json.loads(fixed)
            if isinstance(result, list):
                return result
        except json.JSONDecodeError as e:
            log.debug("json.loads failed after trailing-comma fix: %s", e)

    # ── Step 4: object-by-object extraction ───────────────────────────────────
    # Match {...} blocks that may span multiple lines
    log.debug("Falling back to object-by-object extraction")
    objects = re.findall(r"\{[^{}]+\}", cleaned, re.DOTALL)
    results = []
    for obj_str in objects:
        obj_str = re.sub(r",\s*([\]\}])", r"\1", obj_str).strip()
        try:
            results.append(json.loads(obj_str))
            continue
        except json.JSONDecodeError:
            pass
        # ── Step 5: regex key-value fallback per object ───────────────────
        ind_m  = re.search(r'"indication"\s*:\s*"([^"]*)"', obj_str)
        id_m   = re.search(r'"id"\s*:\s*(?:"([^"]*?)"|null)', obj_str)
        name_m = re.search(r'"name"\s*:\s*(?:"([^"]*?)"|null)', obj_str)
        if ind_m or id_m:
            results.append({
                "indication": ind_m.group(1) if ind_m else "",
                "id":         id_m.group(1)  if (id_m and id_m.group(1)) else None,
                "name":       name_m.group(1) if (name_m and name_m.group(1)) else None,
            })

    if not results:
        log.warning("_clean_and_parse_json: all strategies failed. Raw (500 chars): %s", text[:500])
    return results


def _verify_and_enrich(ind: str, did: str | None, name: str | None) -> tuple[str, str | None, str | None]:
    """Verify disease ID via OT search, fall back to name search if needed."""
    if did:
        did = did.replace(":", "_")
        verified_id, verified_name = ot_search_disease(did)
        if verified_id:
            return ind, verified_id, verified_name
    # Fall back to plain text search
    fb_id, fb_name = ot_search_disease(ind)
    return ind, fb_id, fb_name


def _resolve_single_indication(ind: str) -> tuple[str, str | None, str | None]:
    """Resolve one indication via a dedicated single-item Gemini call."""
    prompt = (
        f"Search OpenTargets (platform.opentargets.org) for the disease that best "
        f"matches the indication \"{ind}\". "
        f'Return ONLY valid JSON (no prose, no fences): {{"indication": "{ind}", "id": "<EFO_/MONDO_/HP_ ID or null>", "name": "<OT disease name or null>"}}'
    )
    text = gemini_search(prompt)
    log.debug("    Single fallback raw: %s", text[:300])
    did, name = None, None
    try:
        objs = _clean_and_parse_json(text)
        if objs:
            obj  = objs[0] if isinstance(objs, list) else objs
            did  = obj.get("id") or None
            name = obj.get("name") or None
    except Exception:
        did = extract_id_from_text(text, "disease")
    return _verify_and_enrich(ind, did, name)


def _resolve_indication_batch(
    batch: list[str],
) -> list[tuple[str, str | None, str | None]]:
    """
    Resolve a batch of up to BATCH_SIZE indications in one Gemini call.
    Returns list of (indication, disease_id, disease_name).
    - Retries the whole batch once on parse failure.
    - Falls back to individual calls if batch still fails.
    """
    numbered = "\n".join(f"{i+1}. {ind}" for i, ind in enumerate(batch))
    prompt = (
        "You are a biomedical data assistant. Search OpenTargets Platform "
        "(platform.opentargets.org) for the best matching disease for each indication below.\n\n"
        "STRICT OUTPUT RULES:\n"
        "- Output ONLY a valid JSON array. Nothing else. No prose. No markdown. No ```json fences.\n"
        "- The array must have EXACTLY one object per indication, in the same order.\n"
        "- Each object must have exactly these three keys:\n"
        '  {"indication": "<copy the indication text exactly>", '
        '"id": "<EFO_XXXXXXX or MONDO_XXXXXXX or HP_XXXXXXX — use underscore not colon>", '
        '"name": "<official OpenTargets disease name>"}\n'
        "- Use JSON null (not the string \"null\") when no match is found.\n"
        "- Do NOT add trailing commas. Do NOT add any text before or after the array.\n\n"
        f"Indications to resolve:\n{numbered}\n\n"
        "JSON array output:"
    )

    def _attempt() -> list[tuple[str, str | None, str | None]]:
        text = gemini_search(prompt)
        log.warning("    Gemini batch raw (first 600 chars):\n%s", text[:600])
        parsed = _clean_and_parse_json(text)
        if not parsed:
            raise ValueError("Empty parse result")
        results = []
        for item in parsed:
            ind  = item.get("indication", "")
            did  = item.get("id") or None
            name = item.get("name") or None
            results.append(_verify_and_enrich(ind, did, name))
        return results

    for attempt in range(1, 3):
        try:
            results = _attempt()
            if len(results) == len(batch):
                return results
            log.warning(
                "Batch attempt %d: got %d/%d results — retrying",
                attempt, len(results), len(batch)
            )
        except Exception as exc:
            log.warning("Batch attempt %d failed (%s) — retrying", attempt, exc)
        time.sleep(2)

    # Fallback: resolve each indication individually
    log.warning("Falling back to per-item resolution for: %s", batch)
    return [_resolve_single_indication(ind) for ind in batch]


def resolve_indications(
    unique_indications: list[str],
    cache: dict,
) -> dict[str, tuple[str | None, str | None]]:
    """
    Resolve indications -> OT diseases.
    - Cache hit with a valid ID: use it directly.
    - Cache hit with null ID (previously unresolved) OR cache miss:
      fall back to ot_search_disease() — a direct OT search API call,
      no Gemini. Updates cache on the fly so subsequent runs benefit.
    OpenTargets scores are always fetched fresh on every run.
    """
    ind_map: dict[str, tuple[str | None, str | None]] = {}
    cached  = cache.get("indications", {})
    cache_updated = False

    for ind in unique_indications:
        entry = cached.get(ind)
        cached_id = entry.get("id") if entry else None

        if entry and cached_id:
            # Good cache hit — use it
            ind_map[ind] = (cached_id, entry.get("name"))
            log.info("  💾 Cache hit '%s' → %s (%s)", ind, cached_id, entry.get("name"))
        else:
            # Cache miss or previously unresolved — query OT search API directly
            if entry:
                log.warning("  ⚠️  Cache has null ID for '%s' — querying OT search API", ind)
            else:
                log.warning("  ⚠️  '%s' not in cache — querying OT search API", ind)

            # Try primary name first, then progressively simplified fallbacks
            ot_id, ot_name = None, None
            search_terms = [ind] + _fallback_search_terms(ind)
            for term in search_terms:
                ot_id, ot_name = ot_search_disease(term)
                if ot_id:
                    if term != ind:
                        log.info("    ✅ OT search resolved '%s' via '%s' → %s (%s)", ind, term, ot_id, ot_name)
                    else:
                        log.info("    ✅ OT search resolved '%s' → %s (%s)", ind, ot_id, ot_name)
                    break
                log.warning("    ↩ Term '%s' → no match", term)

            ind_map[ind] = (ot_id, ot_name)
            cached[ind] = {"id": ot_id, "name": ot_name}
            cache_updated = True

            if not ot_id:
                log.warning("    ❌ OT search could not resolve '%s'", ind)

    if cache_updated:
        cache["indications"] = cached
        save_cache(cache)

    return ind_map


# ── Scoring (parallelised) ─────────────────────────────────────────────────────


def normalize_indication(ind: str) -> str:
    """
    Produce a canonical form so spelling variants like
    'Pre-diabetes', 'Prediabetes', 'pre-diabetes', 'Pre Diabetes'
    all collapse to the same key.

    Steps:
      1. Strip leading/trailing whitespace
      2. Collapse internal whitespace
      3. Lower-case
      4. Remove hyphens between letters (pre-diabetes -> prediabetes)
      5. Collapse any resulting double spaces
    """
    s = ind.strip()
    s = re.sub(r"\s+", " ", s)
    s = s.lower()
    # Remove hyphens between word characters: "pre-diabetes" -> "prediabetes"
    s = re.sub(r"(?<=\w)-(?=\w)", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _fallback_search_terms(ind: str) -> list[str]:
    """
    Generate progressively simpler OT search terms for an indication.
    Applied both at resolution time (cache miss) and scoring time (zero scores).

    Steps applied in order:
      1. Strip parenthetical qualifiers  "Alcohol Use Disorder (AUD)" -> "Alcohol Use Disorder"
      2. Replace separators (/, -)       "Cardiovascular Risk/Disease" -> "Cardiovascular Risk Disease"
      3. Replace hyphens in compound words "Pre-Diabetes" -> "Pre Diabetes"
      4. Progressively shorten: first 3 words, first 2, first word
    Deduplicates and removes the original term.
    """
    terms = []
    # Strip parenthetical acronym/qualifier
    stripped = re.sub(r"\s*\(.*?\)", "", ind).strip()
    if stripped and stripped != ind:
        terms.append(stripped)
    base = stripped if stripped else ind
    # Replace slash separator
    slashed = re.sub(r"[/]", " ", base).strip()
    if slashed != base:
        terms.append(slashed)
        base = slashed
    # Replace hyphen separator (but not at word start like "non-")
    dehyphen = re.sub(r"(?<=[A-Za-z])-(?=[A-Za-z])", " ", base).strip()
    if dehyphen != base:
        terms.append(dehyphen)
        base = dehyphen
    # Shorten progressively
    words = base.split()
    if len(words) > 3:
        terms.append(" ".join(words[:3]))
    if len(words) > 2:
        terms.append(" ".join(words[:2]))
    if len(words) > 1:
        terms.append(words[0])
    # Deduplicate preserving order, skip original
    seen = {ind}
    result = []
    for t in terms:
        if t and t not in seen:
            seen.add(t)
            result.append(t)
    return result


def _extract_gene_search_terms(moa: str) -> list[str]:
    """
    Extract candidate gene/target search terms from a MoA description string.
    MoA strings like "Gastric Inhibitory Polypeptide Receptor Agonist" need to be
    converted to searchable gene names for OT's target search.

    Strategy (in priority order):
      1. Symbol in parentheses: "Calcitonin Receptor (CALCR) Agonist" -> ["CALCR"]
      2. Hyphenated/numeric symbol: "GLP-1 Receptor" -> ["GLP-1"]
      3. All-caps acronym: "GIPR Agonist" -> ["GIPR"]
      4. Strip action word, search protein description:
         "Gastric Inhibitory Polypeptide Receptor Agonist" -> "Gastric Inhibitory Polypeptide Receptor"
    """
    ACTION_WORDS = {"Agonist", "Antagonist", "Activator", "Inhibitor",
                    "Modulator", "Blocker", "Stimulator", "Suppressor"}
    terms = []
    # 1. Symbol in parentheses
    m = re.search(r"\(([A-Z][A-Z0-9\-]+)\)", moa)
    if m:
        terms.append(m.group(1))
    # 2. Hyphenated/numeric acronym e.g. "GLP-1", "IL-6"
    m = re.search(r"\b([A-Z]{2,}[\-]\d+[A-Z]*)\b", moa)
    if m:
        terms.append(m.group(1))
    # 3. All-caps standalone acronym (3+ chars ending in R, P, etc.)
    for word in moa.split():
        if re.fullmatch(r"[A-Z]{3,}\d*", word) and word not in ACTION_WORDS:
            terms.append(word)
            break
    # 4. Strip action word from end and search protein description
    stripped = re.sub(
        r"\s+(" + "|".join(ACTION_WORDS) + r")\s*$", "", moa, flags=re.IGNORECASE
    ).strip()
    stripped = re.sub(r"\s*\(.*?\)", "", stripped).strip()
    if stripped and stripped != moa:
        terms.append(stripped)
    # Deduplicate
    seen: set[str] = set()
    result = []
    for t in terms:
        if t and t not in seen:
            seen.add(t)
            result.append(t)
    return result

def _score_indication(
    ind: str,
    disease_id: str,
    target_ids: list[tuple[str, str, str | None]],
) -> tuple[str, float | None]:
    """
    Score one indication against all MoA targets.

    Uses disease(efoId) -> associatedTargets with Bs (target ID filter) to fetch
    the exact score for each target-disease pair in a single targeted API call.
    This avoids the pagination cut-off of the bulk scan approach: even if a disease
    has thousands of associated targets, passing Bs=[targetId] returns only that
    target's row regardless of its rank.
    """
    wanted = {tid: (moa, sym) for moa, tid, sym in target_ids if tid}
    if not wanted:
        return ind, None

    # One query per target: Bs filters associatedTargets to just that target ID
    # We pass datasources explicitly to match Platform UI weights (API defaults differ).
    query = """
    query TargetDiseaseScore($diseaseId: String!, $targetIds: [String!]!,
                             $datasources: [DatasourceSettingsInput!]) {
      disease(efoId: $diseaseId) {
        associatedTargets(
          enableIndirect: false
          Bs: $targetIds
          datasources: $datasources
          page: { index: 0, size: 1 }
        ) {
          rows {
            score
            target { id approvedSymbol }
          }
        }
      }
    }
    """

    def _fetch_scores_for_disease(did: str) -> list[float]:
        """Fetch scores for all wanted targets against a given disease ID."""
        result = []
        for tid, (moa, sym) in wanted.items():
            data = _ot_post(
                query,
                {"diseaseId": did, "targetIds": [tid],
                 "datasources": OT_DATASOURCE_WEIGHTS},
                context=f"score:{tid}x{did}",
            )
            score = None
            if data:
                rows = (
                    data.get("disease", {})
                        .get("associatedTargets", {})
                        .get("rows", [])
                )
                if rows:
                    score = float(rows[0]["score"])
            log.info("    %s (%s) x '%s' -> %s",
                     sym or moa, tid, ind,
                     f"{score:.4f}" if score is not None else "N/A")
            if score is not None:
                result.append(score)
        return result

    # Primary attempt with the cached disease ID
    scores = _fetch_scores_for_disease(disease_id)

    # If no scores found, try alternative OT disease IDs:
    # 1. Search OT directly with the indication name as-is
    # 2. Try progressively simplified search terms
    if not scores:
        log.warning("  No scores for '%s' with %s — trying fallback disease search", ind, disease_id)
        tried_ids = {disease_id}

        # Build list of search terms to try: full name first, then simplified
        search_terms = [ind] + _fallback_search_terms(ind)

        for term in search_terms:
            alt_id, alt_name = ot_search_disease(term)
            if not alt_id:
                log.warning("  Fallback term '%s' → no OT match", term)
                continue
            if alt_id in tried_ids:
                log.info("  Fallback term '%s' → same ID %s already tried, skipping", term, alt_id)
                continue
            tried_ids.add(alt_id)
            log.info("  Fallback term '%s' → %s (%s) — re-scoring", term, alt_id, alt_name)
            scores = _fetch_scores_for_disease(alt_id)
            if scores:
                break  # found a better ID that yields scores

    best = max(scores) if scores else None
    log.info("  -> Max score for '%s': %s", ind, f"{best:.4f}" if best else "None")
    return ind, best


def compute_max_scores(
    indications: list[str],
    ind_map: dict[str, tuple[str | None, str | None]],
    target_map: dict[str, tuple[str | None, str | None]],
) -> dict[str, float | None]:
    """Fetch association scores for all indications in parallel (WORKERS threads)."""
    target_ids = [
        (moa, tid, sym)
        for moa, (tid, sym) in target_map.items()
        if tid is not None
    ]
    if not target_ids:
        log.error("No valid targets resolved — cannot score.")
        return {ind: None for ind in indications}

    max_scores: dict[str, float | None] = {}
    futures = {}

    with ThreadPoolExecutor(max_workers=WORKERS, thread_name_prefix="score") as exe:
        for ind in indications:
            disease_id, _ = ind_map.get(ind, (None, None))
            if not disease_id:
                log.warning("  Skipping '%s' — no disease ID", ind)
                max_scores[ind] = None
                continue
            log.info("  Queuing '%s' (%s) × %d targets", ind, disease_id, len(target_ids))
            fut = exe.submit(_score_indication, ind, disease_id, target_ids)
            futures[fut] = ind

        for fut in as_completed(futures):
            try:
                ind, best = fut.result()
                max_scores[ind] = best
            except Exception as exc:
                ind = futures[fut]
                log.error("Scoring failed for '%s': %s", ind, exc)
                max_scores[ind] = None

    return max_scores


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    log.info("📂 Reading: %s", INPUT_FILE)
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME or 0, dtype=str)
    df.columns = df.columns.str.strip()

    col_lower = {c.lower(): c for c in df.columns}
    missing   = [c for c in ("indication", "moa") if c not in col_lower]
    if missing:
        sys.exit(f"❌  Missing columns: {missing}. Found: {list(df.columns)}")

    ind_col = col_lower["indication"]
    moa_col = col_lower["moa"]

    raw_moa_str = df[moa_col].dropna().iloc[0]
    raw_moas    = [m.strip() for m in raw_moa_str.split(";") if m.strip()]

    # ── Normalize indications ─────────────────────────────────────────────────
    # Spelling variants like "Pre-diabetes" / "Prediabetes" / "pre-diabetes"
    # must all resolve to the same OT disease ID and share one score.
    # We build a mapping:  normalized_key -> first raw string seen
    # and a reverse map:   raw_string -> normalized_key
    raw_inds_all  = df[ind_col].dropna().unique().tolist()
    norm_to_raw: dict[str, str] = {}   # normalized -> representative raw string
    raw_to_norm: dict[str, str] = {}   # every raw variant -> its normalized key
    for raw in raw_inds_all:
        nk = normalize_indication(raw)
        raw_to_norm[raw] = nk
        if nk not in norm_to_raw:
            norm_to_raw[nk] = raw  # keep the first spelling as representative

    unique_inds = list(norm_to_raw.values())  # one representative per group

    if len(unique_inds) < len(raw_inds_all):
        merged = len(raw_inds_all) - len(unique_inds)
        log.info("🔗 Merged %d spelling variants into canonical indications", merged)

    log.info("🧬 MoAs (%d): %s", len(raw_moas), raw_moas)
    log.info("🦠 Indications (%d): %s", len(unique_inds), unique_inds)
    log.info("🔧 Workers: %d  |  Batch size: %d  |  Cache: %s", WORKERS, BATCH_SIZE, CACHE_FILE)

    cache = load_cache()

    # ── Step 1: MoA → targets (cache-only) ────────────────────────────────────
    log.info("\n── Step 1: Loading MoA targets from cache ────────────────────────")
    target_map = resolve_moa_targets(raw_moas, cache)

    # ── Step 2: Indications → diseases (cache-only) ────────────────────────────
    log.info("\n── Step 2: Loading indications from cache ────────────────────────")
    ind_map = resolve_indications(unique_inds, cache)

    # ── Step 3: Association scores ─────────────────────────────────────────────
    log.info("\n── Step 3: Fetching association scores ───────────────────────────")
    max_scores = compute_max_scores(unique_inds, ind_map, target_map)

    # ── Map results back to every raw indication variant ───────────────────────
    # Each raw string looks up its normalized representative's result
    def _lookup_score(raw_ind):
        rep = norm_to_raw.get(raw_to_norm.get(raw_ind, ""), raw_ind)
        return max_scores.get(rep)

    def _lookup_ind_map(raw_ind):
        rep = norm_to_raw.get(raw_to_norm.get(raw_ind, ""), raw_ind)
        return ind_map.get(rep, (None, None))

    # ── Step 4: Write Excel ────────────────────────────────────────────────────
    df["association_score"]   = df[ind_col].map(_lookup_score)
    df["ot_disease_id"]       = df[ind_col].map(lambda x: _lookup_ind_map(x)[0])
    df["ot_disease_name"]     = df[ind_col].map(lambda x: _lookup_ind_map(x)[1])
    df["ot_targets_resolved"] = "; ".join(
        f"{moa}→{sym or tid}"
        for moa, (tid, sym) in target_map.items() if tid
    )

    log.info("\n── Step 4: Saving to %s", OUTPUT_FILE)
    from openpyxl.styles import Font, PatternFill, Alignment
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")
        ws = writer.sheets["Results"]

        hdr_fill   = PatternFill("solid", fgColor="4F81BD")
        score_fill = PatternFill("solid", fgColor="E2EFDA")

        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        score_col = df.columns.get_loc("association_score") + 1
        for row in ws.iter_rows(min_row=2, min_col=score_col, max_col=score_col):
            for cell in row:
                cell.fill         = score_fill
                cell.font         = Font(bold=True, name="Arial")
                cell.number_format = "0.0000"

        for col in ws.columns:
            width = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 4, 50)

    log.info("✅ Saved: %s", OUTPUT_FILE)

    # ── Summary ────────────────────────────────────────────────────────────────
    n_targets = sum(1 for v in target_map.values() if v[0])
    n_ind     = sum(1 for v in ind_map.values() if v[0])
    n_scored  = sum(1 for v in max_scores.values() if v is not None)

    log.info("\n── Summary ───────────────────────────────────────────────────────")
    log.info("  MoA targets resolved : %d / %d", n_targets, len(raw_moas))
    log.info("  Indications resolved : %d / %d", n_ind, len(unique_inds))
    log.info("  Scores found         : %d / %d", n_scored, len(unique_inds))
    log.info("  Cache saved to       : %s", CACHE_FILE)

    print("\n📊 Score results:")
    for ind in unique_inds:
        score = max_scores.get(ind)
        did, dname = ind_map.get(ind, (None, None))
        score_str = f"{score:.4f}" if score is not None else "N/A"
        # Show merged variants if any
        variants = [r for r, nk in raw_to_norm.items()
                    if norm_to_raw[nk] == ind and r != ind]
        suffix = f"  (also: {', '.join(variants)})" if variants else ""
        print(f"  {ind:<45} {score_str:<10}  ({dname or did or 'unresolved'}){suffix}")


if __name__ == "__main__":
    main()
